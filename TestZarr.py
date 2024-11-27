from openpyxl import load_workbook
import time
import zarr
from numcodecs import Blosc
from matplotlib import pyplot as plt
import psutil
import time
import resource
import numpy as np

workbook = load_workbook(filename="/Users/sagnik/Documents/Research_Project/Data/Settings.xlsx")


compressor_sheet = workbook["Compressor"]
query_sheet = workbook["Query"]
arr_read_time = []
arr_write_time = []

for index, value in enumerate(compressor_sheet.iter_rows(min_row=2)):
    #Initializing the compressor parameter and configuring it to pass that in the Zarr creation parameter
    compressor = Blosc(cname=value[0].value, clevel=int(value[1].value), shuffle=int(value[2].value))
    zarr_blosc = zarr.create((1000, 1000, 1000), chunks=True, compressor=compressor, dtype='i4')
    zarr_blosc = zarr.array(np.arange(1000000000).reshape(1000, 1000, 1000))
    #start_time = time.perf_counter()
    # Start measuring CPU usage
    start_cpu_usage = psutil.cpu_percent()
    start_time = time.time()
    #Initial Disk I/O usage
    start_reads = psutil.disk_io_counters().read_count
    # start_rusage = resource.getrusage(resource.RUSAGE_SELF)
    # start_reads = start_rusage.ru_disk_reads
    #reading the whole data
    zarr_blosc[:]
    end_cpu_usage = psutil.cpu_percent()
    end_time = time.time()
    # Get final disk I/O usage
    end_reads = psutil.disk_io_counters().read_count
    # end_rusage = resource.getrusage(resource.RUSAGE_SELF)
    # end_reads = end_rusage.ru_disk_reads
    # Calculate the average CPU usage
    total_time = end_time - start_time
    average_cpu_usage = (end_cpu_usage - start_cpu_usage)
    #end_time = time.perf_counter()
    read_time = end_time - start_time
    #Writing the read time in the excel row.
    compressor_sheet.cell(row=index+2, column=5).value = total_time
    print("Read Time: ",end_time - start_time, "seconds")
    #start_time = time.perf_counter()
    #Writing the CPU usage in the CPU_Usage column
    compressor_sheet.cell(row=index+2, column=6).value = average_cpu_usage
    reads = end_reads - start_reads #This is the Disk Reads
    compressor_sheet.cell(row=index+2, column=7).value = reads
    #Writing the whole data
    # zarr_blosc[:] = 1
    # end_time = time.perf_counter()
    # write_time = end_time - start_time
    # compressor_sheet.cell(row=index+2, column=6).value = write_time
    #Writing the write time in the excel row
    #print("Write Time: ",end_time - start_time, "seconds") 
    arr_read_time.append(read_time)
    #arr_write_time.append(write_time)
    
workbook.save("/Users/sagnik/Documents/Research_Project/Data/Settings.xlsx")
workbook.close()


